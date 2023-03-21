'''
    Script created by adres433

    Skrypt na podstawie dostarczonych danych przez użytkownika
    generuje harmonogram pracy dla określonej ilości pracowników
    z podziałem na max. 3 zmiany trwające max. 5 dni w określonym miesiącu
    wskazanego roku.
    Wygenerowany harmonogram wyświetlany jest do wglądu w konsoli oraz
    zapisywany jest do pliku *.xlsx z formatowaniem zawartym w pliku sample.xlsx.
'''

from datetime import date, timedelta
from openpyxl import load_workbook

wb = load_workbook('sample-nie_usuwaj.xlsx')
sheet = wb.active
choose_date = date(2021,5,30)       #rozp. prac nad skryptem
today = date.today()

print("Witaj w generatorze planów pracy. \n")
print("Dzisiaj jest: ", today, '\n\n')

def get_data():
    try:
        month = input("Wybierz miesiąc [MM]:")
        if month == '':
            month = int(today.month)
            print(month)
        else:
            month = int(month)
            if month > 12 or month < 1:
                raise ValueError('Błędny przedział liczbowy.')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość [MM][przedział: 1-12].")
        return False
    try:
        year = input("Wybierz rok [yyyy]:")
        if year == '':
            year = int(today.year)
            print(year)
        else:
            year = int(year)
            if year < 2000:
                raise ValueError('Błędny przedział liczbowy. [>1999]')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość [YYYY].")
        return False
    try:
        worker = int(input("Ilu pracowników [przedział: 1-25]:"))
        if worker > 25 or worker < 1:
            raise ValueError('Błędny przedział liczbowy.')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość.")
        return False
    try:
        seasons = int(input("Ile zmian [1-3]:"))
        if seasons > 3 or seasons < 1:
            raise ValueError('Błędny przedział liczbowy.')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość [przedział: 1-3].")
        return False
    try:
        shift = int(input("Dni na zmianę [1-5]:"))
        if shift > 5 or shift < 1:
            raise ValueError('Błędny przedział liczbowy.')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość [przedział: 1-5].")
        return False
    try:
        first = int(input("Pierwsza zmiana pierwszego dnia - który pracownik [1-"+str(worker)+"]:"))
        if first > worker or first < 1:
            raise ValueError('Błędny przedział liczbowy.')
    except ValueError:
        print("\n\nPodałeś nieprawidłową wartość [przedział: 1-5].")
        return False
    return [year, month, worker, seasons, shift, first]

def generate():

    global choose_date
    list = get_data()
    sheet["A40"].value = str(list)
    sheet["A41"].value = "© by adres433"
    try:
        if list[1] < 12:
            choose_date = date(list[0],list[1]+1, 1)            #wybrany miesiąc +1
        else:
            choose_date = date(list[0]+1,1, 1)                    #wybrany miesiąc +1
    except:
        print("Błąd podczas generowania.")
        return 0x16
    how_day = choose_date - timedelta(days=1)
    choose_date = date(list[0],list[1], 1)              #wybrany miesiąc
    how_day = how_day.day                               #ilosć dni w miesiącu
    month = []
    month.append([])
    for x in range(how_day):                            #lista dni w miesiącu
        if x+1 < 10:
            month[0].append('0'+str(x+1))               #z bierzącym 0
        else:
            month[0].append(str(x+1))                   #bez zera
    for y in range(list[3]):                            #ilość zmian
        temp = []
        for x in range(how_day):                        #dla każdego dnia w miesiącu dla danej zmiany
            temp.append(str(list[5]))
            if (x+1)%list[4] == 0:
                list[5] = int(list[5])+1
                if list[5] > list[2]:
                    list[5] = 1
        month.append(temp)
        if list[5] == 0:
            list[5] = list[2]
    return [month, list[2]]

def format():
    data = generate()
    try:
        print(data[0][0])
    except:
        print("Błąd podczas formatowania wyników. \nSpróbuj ponownie :)\n\n")
        if data == 0x16:
            return format()
        else:
            return
    workers = []
    for x in range(data[1]):    #dla każdego pracownika
        workers.append([])
        for y in range(len(data[0])-1): #dla każdej zmiany
            for index, z in enumerate(data[0][y+1]):    #dla każdego dnia miesiąca
                if int(z) == x+1:
                    t = ''
                    if y+1 == 1:
                        t = sheet['X32'].value
                    elif y+1 == 2:
                        t = sheet['X36'].value
                    else:
                        t = sheet['X34'].value
                    sheet.cell(row=2+x, column=2+index).value = t     #dodanie informacji o zmianie aktualnego pracownika do arkusza excel
                    if y == 0:
                        if y+1 < 10:
                            workers[x].append(" "+t)
                        else:
                            workers[x].append(t)
                    else:
                        if y+1 < 10:
                            workers[x][index] = " "+t
                        else:
                            workers[x][index] = t
                else:
                    if y == 0:
                        workers[x].append('  ')
        if x != 0:
            print(data[0][0])
        print(workers[x])
        print(sheet["A"+str(x+2)].value, end=' ')               #nazwa pracownika z arkusz sample.xlsx
        print("- przepracowanych dni: ",str(len(workers[x])-workers[x].count('  ')))
        sheet["A28"].value = "=UPPER(TEXT(\""+str(choose_date)+"\",\"mmmm\"))"
        sheet["A32"].value = str(choose_date.year)
        print('\n')
    return

format()
wb.save("plan_"+str(choose_date.month)+"_"+str(choose_date.year)+".xlsx")
input("Naciśnij ENTER, aby zamknąć.")